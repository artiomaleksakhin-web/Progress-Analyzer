"""
Анализатор успеваемости студентов.

Программа загружает данные об оценках студентов из CSV или Excel-файла,
рассчитывает средний балл, определяет статус успеваемости и выводит
общую статистику по группе.
"""

import csv
from pathlib import Path
from typing import Dict, List

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

Student = Dict[str, object]
AnalysisRow = Dict[str, object]

SUBJECTS = ["Математика", "Русский язык", "Информатика", "История"]


def _validate_grade(value: object, student_name: str, subject: str) -> int:
    """Проверяет оценку и возвращает ее как целое число."""
    try:
        grade = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"У студента {student_name!r} по предмету {subject!r} указана "
            f"некорректная оценка: {value!r}"
        ) from exc

    if grade < 2 or grade > 5:
        raise ValueError(
            f"Оценка должна быть от 2 до 5. Студент: {student_name}, "
            f"предмет: {subject}, значение: {grade}"
        )

    return grade


def load_students_from_csv(file_path: str) -> List[Student]:
    """Загружает список студентов из CSV-файла."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {file_path}")

    students: List[Student] = []
    required_columns = ["ФИО"] + SUBJECTS

    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file, delimiter=";")

        if reader.fieldnames is None:
            raise ValueError("CSV-файл пустой или не содержит заголовков")

        missing_columns = [col for col in required_columns if col not in reader.fieldnames]
        if missing_columns:
            raise ValueError(
                "В CSV-файле отсутствуют обязательные столбцы: "
                + ", ".join(missing_columns)
            )

        for row in reader:
            name = str(row["ФИО"]).strip()
            if not name:
                continue

            grades = {
                subject: _validate_grade(row[subject], name, subject)
                for subject in SUBJECTS
            }
            students.append({"ФИО": name, "Оценки": grades})

    return students


def load_students_from_excel(file_path: str) -> List[Student]:
    """Загружает список студентов из Excel-файла."""
    if load_workbook is None:
        raise ImportError("Установите библиотеку openpyxl: pip install openpyxl")

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {file_path}")

    workbook = load_workbook(path)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    required_columns = ["ФИО"] + SUBJECTS
    missing_columns = [col for col in required_columns if col not in headers]
    if missing_columns:
        raise ValueError(
            "В Excel-файле отсутствуют обязательные столбцы: "
            + ", ".join(missing_columns)
        )

    column_indexes = {header: headers.index(header) for header in required_columns}
    students: List[Student] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name_cell = row[column_indexes["ФИО"]]
        if name_cell is None or str(name_cell).strip() == "":
            continue

        name = str(name_cell).strip()
        grades = {
            subject: _validate_grade(row[column_indexes[subject]], name, subject)
            for subject in SUBJECTS
        }
        students.append({"ФИО": name, "Оценки": grades})

    return students


def calculate_average(grades: Dict[str, int]) -> float:
    """Рассчитывает средний балл студента."""
    return round(sum(grades.values()) / len(grades), 2)


def define_status(average_grade: float) -> str:
    """Определяет статус успеваемости по среднему баллу."""
    if average_grade >= 4.5:
        return "отличник"
    if average_grade >= 3.5:
        return "хорошист"
    if average_grade >= 2.5:
        return "троечник"
    return "неуспевающий"


def analyze_students(students: List[Student]) -> List[AnalysisRow]:
    """Формирует итоговый анализ по каждому студенту."""
    result: List[AnalysisRow] = []

    for student in students:
        name = str(student["ФИО"])
        grades = student["Оценки"]
        if not isinstance(grades, dict):
            raise TypeError("Оценки студента должны быть словарем")

        average = calculate_average(grades)
        result.append({
            "ФИО": name,
            "Средний балл": average,
            "Статус": define_status(average)
        })

    return result


def calculate_group_average(analysis: List[AnalysisRow]) -> float:
    """Рассчитывает средний балл группы."""
    if not analysis:
        return 0.0
    total = sum(float(row["Средний балл"]) for row in analysis)
    return round(total / len(analysis), 2)


def find_best_student(analysis: List[AnalysisRow]) -> AnalysisRow:
    """Находит студента с самым высоким средним баллом."""
    return max(analysis, key=lambda row: float(row["Средний балл"]))


def find_weak_student(analysis: List[AnalysisRow]) -> AnalysisRow:
    """Находит студента с самым низким средним баллом."""
    return min(analysis, key=lambda row: float(row["Средний балл"]))


def print_report(analysis: List[AnalysisRow]) -> None:
    """Выводит результат анализа в консоль."""
    if not analysis:
        print("Нет данных для анализа.")
        return

    print("АНАЛИЗАТОР УСПЕВАЕМОСТИ")
    print("-" * 70)
    print(f"{'ФИО':<30} {'Средний балл':<15} {'Статус'}")
    print("-" * 70)

    for row in analysis:
        print(f"{row['ФИО']:<30} {row['Средний балл']:<15} {row['Статус']}")

    print("-" * 70)
    print(f"Средний балл группы: {calculate_group_average(analysis)}")

    best_student = find_best_student(analysis)
    weak_student = find_weak_student(analysis)

    print(f"Лучший студент: {best_student['ФИО']} ({best_student['Средний балл']})")
    print(f"Самый низкий средний балл: {weak_student['ФИО']} ({weak_student['Средний балл']})")


def main() -> None:
    students = load_students_from_csv("students.csv")
    # Для чтения Excel-файла можно использовать строку ниже:
    # students = load_students_from_excel("students.xlsx")
    analysis = analyze_students(students)
    print_report(analysis)


if __name__ == "__main__":
    main()
